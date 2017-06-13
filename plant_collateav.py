#!usr/bin/env python

"""Program to process scoring data"""
#Written by Bhupinder Sehra; created 8/29/16

import sys
import os
import pandas as pd
import numpy as np
import itertools
import openpyxl
import collections

xfin = sys.argv[1]
sheetin = str(sys.argv[2])
devstagenum = sys.argv[3]
xfout = sys.argv[4]
xsheetout = sys.argv[5]
newwb = str(sys.argv[6])

#numtissues = sys.argv[4]

def getcol(matrix, i):
    return [row[i] for row in matrix]

def getavscores(devstagelist, onesiblist, stagecolnum):
    avscorelist = [] #will contain average scores                                                                                                                                                                     
    templ = []
    collate = []
    zcollate = []
    avs = []
    newlist = []
    denom = 0
    counter = 0
    for stg in devstagelist:
        for srow in onesiblist:
            if str(srow[stagecolnum]) == str(stg):
                collate.append(srow[3:])
                #print "This is srow", srow[3:]
            #avs.append(srow[:2])
        for index in xrange(2):
            avs.append(srow[index])
        avs.append(stg)
        #print avs, "This is avs"
        zcollate = zip(*collate)
            #print "This is zcollate", zcollate
        #print zcollate[0], "This is zcollate[0]"
        #dealing with style expression (y/n values)
        for i in xrange(len(zcollate)):
                #print "row in zcollate", zcollate
            if 'n' in zcollate[i] or 'y' in zcollate[i]:
                if 'y' in zcollate[i]:
                    avs.append("y")
                else:
                    avs.append("n")
            else:
                avs.append(float(sum([fl for fl in zcollate[i]]))/float(len(zcollate[i])))
                #print "This is what was added:", "sum", float(sum([fl for fl in zcollate[i]])), "len",  float(len(zcollate[i])), "sum/len", float(sum([fl for fl in zcollate[i]]))/float(len(zcollate[i]))
        newlist.append(avs)
            #print newlist, "This is newlist"
        avs = []
        collate = []
    return newlist

#declaring variables
df = {}
df2 = {}
sorteddf = {}
stagelist = []
labels = []
vsetsib = []
usiblist = []
sibrows = []
scorerows = []
finalrows = []
rcol = []
count = []
tissues=[]
stagecount = []
nptissuecount = []
sctrans = []
sctlen = 0
numrows= 0
numcols = 0
sibav =[]
allsibav = []
df = pd.read_excel(xfin, sheetname=None)
df2 = df.get(sheetin)
#print df2
tempdf = {}

for i in range(0, int(devstagenum)):  
    stagelist.append(str(df2.iloc[i, 2]))
tissues = df2.columns.values.tolist()
print "This is tissues", tissues
setsib = set(df2.iloc[:, 1].values.tolist())
setsib = set(filter(lambda x: x == x , setsib))
temp =[]
#collect alike sibling rows in main df
for sib in setsib:
   print "dealing with sib ", sib
   for row in df2.itertuples():
       try:
           if str(row[2]) == "":
               continue
           if row[2] == sib:
               sibrows.append(list(row)[1:])
       except:
           print "what you have cannot be converted to float"
   #print len(sibrows), "This is the length of sibrows"   
   if len(sibrows) == 7:
       tempdf = pd.DataFrame(data=sibrows, columns=tissues)
       tempdf.set_value(0, 'line', sib)
       print "This is tempdf for sibs with only 1 plant", tempdf
       #sibrows = []
       #dfx = tempdf.loc[:,0]
       #allsibav.append(tempdf)
   if len(sibrows) > 7:
       print "yes more than 7", sibrows
       sibav = getavscores(stagelist, sibrows, 2)
       tempdf = pd.DataFrame(data=sibav, columns=tissues)
       tempdf.set_value(0, 'line', sib)
       print "This is tempdf for sibs with > 1 plant", tempdf
   allsibav.append(tempdf)
   sibrows = []
   print len(allsibav), "This is len of allsibav"
#print setsib

#print allsibav, "This is allsibav"
#print allsibsplusav, "this is the latest allsibsplusav"
sorteddf = pd.concat(allsibav, ignore_index=False)
#sorteddf = sorteddf.ix[:,1:]

#concatdf = pd.DataFrame(data=dataout, columns=tissues)
#sorteddf = pd.concat(allsibsplusav, keys=)
#print sorteddf, "this is sorteddf" #write out to Excel

#concatdft.to_excel('test.xlsx', sheet_name='2082_2081_1kbpSHP2_counts')
#concatdf.transpose().to_excel(xfout, sheet_name=xsheetout)

"""
from openpyxl import load_workbook
book = load_workbook(xfout)
writer = pd.ExcelWriter(xfout, engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
sorteddf.to_excel(writer, xsheetout)
writer.save()
"""

if newwb == 'n':
    from openpyxl import load_workbook
    book = load_workbook(xfout)
    writer = pd.ExcelWriter(xfout, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    sorteddf.to_excel(writer, xsheetout)
    writer.save()

if newwb == "y":
    sorteddf.to_excel(xfout, sheet_name=xsheetout)














